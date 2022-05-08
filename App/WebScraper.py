from datetime import date, datetime
import math
from pickletools import long1
from urllib import request as _requests
from openpyxl import Workbook
import requests
import bs4
from numpy import NaN
import requests
import yfinance
import pandas
import dataHandler
import CsvParser
import bs4 as bs #beautiful soup
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

# import aidansThing

# a = yfinance.Ticker("A")
sharpes = pandas.DataFrame()
output_monthlyData= pandas.DataFrame()
period = "5y"
interval = '1mo'
interval = '1d'


def getPriceUI(ticker):
    print("getting prices")
    user_agent_headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
    BASE_URL = str("https://finance.yahoo.com/quote/")
    url_string = str(BASE_URL)
    qs = "?p="
    queryString = str(qs)
    yahoo_URL = url_string + ticker + queryString + ticker
    html = requests.get(yahoo_URL)

    # print("yahoo url = "+yahoo_URL)

    # print("\n HTML = "+str(html))

    if(str(html)=='<Response [404]>'):
        print("HTML request failed to connect, trying with header!")
        html = requests.get(url = yahoo_URL,headers=user_agent_headers)

    soup = bs.BeautifulSoup(html.text,"lxml")

    divName = soup.find('div',{'class':'D(ib) Mt(-5px) Maw(38%)--tab768 Maw(38%) Mend(10px) Ov(h) smartphone_Maw(85%) smartphone_Mend(0px)'})
    timeStamp = soup.find('div',{'class':'C($tertiaryColor) D(b) Fz(12px) Fw(n) Mstart(0)--mobpsm Mt(6px)--mobpsm Whs(n)'})
    modified_name = soup.find('h1',{'class':'D(ib) Fz(18px)'})

    n = str(modified_name).replace('h1','h4')
    t = str(timeStamp)
    divPrice = soup.find('div',{'class':'D(ib) Mend(20px)'})
    # print(divPrice)
    p = str(divPrice) #find value = ""
    # print(p)
    print("\n\n\n")
    p2 = p.split('value="')
    # print(p2)
    print("\n\n\n")

    if(len(p2)>1):
        # print(p2[1])
        i = p2[1].find('"')
        p2 = p2[1]
    else:
        # print("\n\n\n\n P2 WTF LOL ;KLASJDFJKLADHSLJ\n\n"+str(p2)+"\n len of string ="+str(len(p2)))
        i=str(p2).find('"')
    
    p = p2[0:i]

    return n,divPrice,p,t
    

def getHist(ticker,interval):
    print('Ticker = ', ticker)
    print('Interval = ', interval)
    a = yfinance.Ticker(str(ticker))
    print("request = ",a)
    hist = a.history(period,interval)
    print("ARE THERE NANS IN HISTORICAL DATA", (NaN in hist))
    print("hist =",hist)
    if(interval == '1d'):
        df = dataHandler.getData(hist)
        sharpe = CsvParser.calculateStuff(df,ticker)[0]
    else:
        out = CsvParser.calculateStuff(hist,ticker)
        mData = out[1]
        sharpe = out[0]
    print("sharpe = ",sharpe)

    row = pandas.DataFrame({
        "Ticker":[ticker],
        "Sharpe":[sharpe]
        })

    sharpes = pandas.DataFrame()

    print("\n SHARPES?? = "+sharpes)

    sharpes = pandas.concat(
        [sharpes,row],
        ignore_index=True
        ).reset_index(drop=True)

    return sharpe ,mData

def multipleTickers(tickers,interval,loadingBar,root):
    sharpes = pandas.DataFrame()
    monthData = pandas.DataFrame()
    # tickers = ['A' ,'B']
    for ticker in tickers:
        print("=====================",ticker,"==========================")
        a = yfinance.Ticker(str(ticker))
        hist = a.history(period,interval)
        df = dataHandler.getData(hist)
        if(interval == '1d'):
            df = dataHandler.getData(hist)[0]
            sharpe = CsvParser.calculateStuff(df,ticker)[0]
        else:
            sharpe = CsvParser.calculateStuff(hist,ticker)[0]
            moData = CsvParser.calculateStuff(hist,ticker)[1]

        row = pandas.DataFrame({
            "Ticker":[ticker],
            "Sharpe":[sharpe],
            "Modifed Sharpe": [(sharpe * math.sqrt(12))]
        })

        moData2 = CsvParser.correlationTool(ticker)
        print(moData2)
        monthData = pandas.concat(
            [moData2]
        )

        sharpes = pandas.concat(
            [sharpes,row],
            ignore_index=True
            ).reset_index(drop=True)

        loadingBar.step()
        root.update()
        # loadingBar.
        
        # sharpes.append({"Ticker":ticker,"Sharpe":[sharpe]},ignore_index=True)

        # row = pandas.DataFrame({
        #     "Ticker": ticker,
        #     "Sharpe":float(sharpe)
        # })
        # sharpes.append({"Ticker":ticker,"Sharpe":[sharpe]},ignore_index=True)
    print("\n\n\n MonthlyData from run all = ",monthData)
    output_monthlyData = monthData
    runCorrelation(sharpes,monthData)

    return sharpes, monthData

def createMegaDF(tickers):
    MegaDf = pandas.DataFrame()
    ticker_monthlyData = []
    monthlyData = pandas.DataFrame()
    for ticker in tickers:
        ticker_monthlyData = []
        # val = pandas.DataFrame(CsvParser.correlationTool(ticker))
        # MegaDf.insert(0,ticker,val)
        file = "Ticker Data/"+ticker+".xlsx"
        print("\nfile = "+file)
        workBook = openpyxl.load_workbook(file)
        print("workbook = "+str(workBook))
        workSheet = workBook["Monthly Data"]
        # workSheet.values
        monthlyDataCol = workSheet['B']
        # monthlyDataColval = monthlyDataCol[0]
        print("column length =" +str(len(monthlyDataCol)))

        for v in workSheet.values:
            # print("value = "+str(v[1])+",\n \t type: "+str(type(v[1])))
            if(type(v[1]) == float):
                print(" adding "+str(v[1]))
                ticker_monthlyData.append(v[1])
            elif(type(v[1]) == int):
                print(" adding "+str(v[1]))
                ticker_monthlyData.append(v[1])
            else:
                print("none or string = "+str(type(v[1])))
                
        print("\nticker "+str(ticker)+"\n \t data len " +str(len(ticker_monthlyData)))
        print("\n --------\n data:\n",ticker_monthlyData)

        if(len(ticker_monthlyData)!= 61):
            print(" skipping this iteration! not enough data to correlate! ")
            continue
        
        
        # for val in monthlyDataCol:
        #     if(type(val.value)==float):
        #         ticker_monthlyData.append(val.value)
        
        # MegaDf[ticker] = pandas.Series(ticker_monthlyData)
        # MegaDf.concat[pandas.Series(ticker_monthlyData)]
        # MegaDf.insert(0,ticker,ticker_monthlyData)
        MegaDf[ticker] = ticker_monthlyData
        print("\n MEGA DF\n",MegaDf[ticker])
        # print("\n monthly data from "+ticker+" is \n first val = "+str(ticker_monthlyData[0]))
        # MegaDf.insert(0,ticker,monthlyData)
    
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(MegaDf,index=True,header=True):
        ws.append(r)
    # this should have all the ticker monthly data
    ws_corr = wb.create_sheet("coorelation")
    corr_MegaDf=MegaDf.corr()
    for r in dataframe_to_rows(corr_MegaDf,index=True,header=True):
        ws_corr.append(r)

    wb.save("MEGA.xlsx")
    print("Finished MEGA Created")
        
    


def runCorrelation(sharpes,monthData):
    wb = Workbook()
    Sharpe = wb.active
    Sharpe.title= "Sharpes"
    for r in dataframe_to_rows(sharpes,index=True,header=True):
        Sharpe.append(r) 

    MonthlyData = wb.create_sheet("Total Monthly Data")
    for r in dataframe_to_rows(monthData,index=True,header=True):
        MonthlyData.append(r)
    
    correlation = monthData.corr()
    Correlation = wb.create_sheet("Correlation")
    for r in dataframe_to_rows(correlation,index=True,header=True):
        Correlation.append(r)
    
    # sharpes.to_csv("Ticker Data/_TickersNSharpes.csv")
    print("Added Tickers and Sharpes")
    wb.save("Ticker Data/_MainInfo.xlsx")
    print('COMPLETELY FINISHED')

# def postProcessing(output):
#     print("SHARPES +++++++++++++++++++++++++++++++++ \n\n\n\n\n")
#     print(output,range(output))
#     with open("Tickers/TickersNSharpes.csv", "w",newline="") as f:
#         writer = csv.writer(f)
#         reader = csv.reader(f)
        


# getHist("BRK.B",'1mo')
# x = multipleTickers(tickers,'1mo')
# print(x)