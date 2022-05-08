import csv , itertools , operator
from fileinput import filename
from datetime import datetime
import os
import statistics
from tkinter import END
from numpy import NaN
import openpyxl
import pandas
# import openpyxl
from openpyxl import *
from openpyxl.utils.dataframe import dataframe_to_rows
import WebScraper

# class CsvManager:
#     def __init__(self,filename):
#         self.filename = filename

def calculateStuff(dataFrame,ticker):
    print("[CSV PARSER]")
    path =  os.path.realpath(__file__)
    dataFrame = dataFrame.dropna()

    monthlyValues = dataFrame.get('Adj Close')

    # dataFrame.to_csv("Tickers/"+ticker+".csv")
    
    # print("\n",dataFrame,"\n\n\n\n\n\n")
    # print("=======================\n MONTHLY VALUES",monthlyValues,"\n===================")
    l = monthlyValues.size
    monthlyReturns = []
    prevVal = 0
    price = WebScraper.getPriceUI(ticker)
    # print("LENGTH OF MONTHLY VALUE",l)

    for index in monthlyValues:
        # print(index)
        # print("[INDEX ]",type(index))

        
        if index == monthlyValues[0]:
            prevVal = index
            pass
        elif index == 0:
            prevVal = index
            pass
        else:
            monthlyReturns.append(index / prevVal - 1)
            prevVal = index
        if l -1 == index:
            break

    riskRate = 0.0009
    # try: 
    print("MEAN ; ", statistics.mean(monthlyReturns))
    print("standard deviation: ",statistics.stdev(monthlyReturns))
    try: 
        SharpeRate = (statistics.mean(monthlyReturns) - riskRate )/ statistics.stdev(monthlyReturns)
    except:
        print("Issue with calculation... ticker may be wrong or delisted. ")
    # except: 
        # SharpeRate = NaN
        # print("Ticker may be delisted")

    # print("[CSV PARSER FINISHED] SharpeRate = ",SharpeRate)

    if not os.path.exists('Ticker Data'):
        print("NO FILE FOUND CREATING DIRECTORY")
        os.makedirs('Ticker Data')
    else:
        print("file found "+os.path.dirname(os.path.abspath('Ticker Data'))+" ... directory not made")

    # df = pandas.DataFrame(SharpeRate,columns='Sharpe for '+ticker)
    # print("new file created with Sharpe: Tickers/"+ticker+".csv")
    # c = df.to_csv("Tickers/"+ticker+".csv")
    sharpeDF = pandas.DataFrame([ticker,SharpeRate])
    # sharpeDF.to_csv("Tickers/"+ticker+"_Sharpe.csv")
    # dataFrame.to_csv("Tickers Data/"+ticker+"_MonthlyData"+".csv")
    """     dataFrame
    p = "Tickers/"+ticker+".csv"
    f = open(p, 'w')
    reader = csv.reader(f)
    writer = csv.writer(f)
    for row in reader:
        rows = row.copy()
        print(row)
        for i in range(rows)+1:
            rows.append()
    writer.writerow()
    writer.writerow(END,["SharpeRate", SharpeRate, SharpeRate, SharpeRate, SharpeRate])
    f.close()
    fpath = "Tickers/"+ticker+".csv"
    with open (fpath,'w',encoding='UTF8') as f:
        writer = csv.writer(f)
        writer.writerows(SharpeRate) """
    # print("\n\nPRINTING ROWS:\n")
    # print(monthlyValues)

    mVals = pandas.DataFrame(monthlyValues)
    wb = Workbook()
    # ws = wb.active
    rows = dataframe_to_rows(monthlyValues)
    # monthly_Data = wb.create_sheet("Monthly Data")
    monthly_Data = wb.active
    monthly_Data.title = "Monthly Data"
    monthly_Data['A1'] = "Monthly Data"

    for r in dataframe_to_rows(mVals, index=True, header=True):
        monthly_Data.append(r)


    sharpe_Data = wb.create_sheet("Sharpe Data")
    sharpe_Data['A1'] = "Sharpe Data"

    for r in dataframe_to_rows(sharpeDF, index=True, header=True):
        sharpe_Data.append(r)


    # print(str(datetime.today))
    # print(datetime.today())
    price_Data = wb.create_sheet("price")
    price_Data['A1'] = "Price"
    price_Data['A2'] = price[2] 
    price_Data['B1'] = 'Date'
    price_Data['B2'] = datetime.today()
    # _excel.save(ticker+".xlsx")
    # print("rows type = ",type(rows))
    # print("monthlyvalues dataframe type = ",type(monthlyValues))
    # print("mVals dataframe type = ",type(mVals))

    wb.save("Ticker Data/"+ticker+".xlsx")

    # _excel = openpyxl.Workbook()
    # excel = _excel.active 

    # p = dataframe_to_rows(dataFrame)
    # print(p)



    return SharpeRate , monthlyValues

def correlationTool(ticker):
     
    #need to select multiple stocks ... monthly returns for each 
    #ideally all the stocks...
    # numpy.corrcoef(monthylDataStock1,monthylDataStock2) 
    #correllation matrix tool  
    #   
    # just make a dataframe from the excel sheet with all of the monthly data 
    # do df.corr()
    
    def getTicker1moData():
        print("getting ticker 1moData")
        file = "Ticker Data/"+ticker +".xlsx"
        wb = load_workbook(str(file))
        # print(wb.sheetnames)
        ws = wb["Monthly Data"]
        colB = ws['B']

        moData = []
        for val in colB:
            # print(val.value)
            # print(type(val.value))
            if(type(val.value)==float):
                moData.append(val.value)
            
           
        # print(moData)
        s = pandas.Series(moData)
        df2 = pandas.DataFrame()
        df2[ticker] = pandas.Series(moData)
        return df2

        
    return getTicker1moData()
        
    
    # for ticker in tickers:
    #     excelName = ticker+".xlsx"
    #     moData = getTicker1MoData(excelName) # gets monthly data from excel
    #     tickers_new = tickers.split(ticker) # ticker should always be the first one
    #     for ticker2 in tickers_new:
    #         #ticker2 is now the second value within the thing.
    #          excelName_new = ticker+".xlsx"
    #          moData_new = getTicker1MoData(excelName_new) # gets monthly data from excel
    #          numpy.corrcoef(moData,moData_new)

        

